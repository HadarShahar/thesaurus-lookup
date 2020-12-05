"""
    Hadar Shahar

    NOTE: thesaurus_threads.py does the same thing, but obviously much faster.

    Save all your words in a file at WORDS_LIST_PATH,
    and get an excel sheet with synonyms and sentences for each word.
"""
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import re
import os

SYNONYMS_NUM = 3  # synonyms per word
SENTENCES_NUM = 1  # sentences per word
WORDS_LIST_PATH = 'words_list.txt'  # the input file
OUTPUT_FILENAME = 'synonyms and sentences.xlsx'
FONT_SIZE = 16
WORDS_SEPARATORS = [',', '/']

# necessary for finding the html elements at thesaurus.com
# the class of the synonyms divs in thesaurus.com
SYNONYMS_DIV_CLASS = 'css-1s9gh2j etbu2a30'
# the class of the div that contains the EXAMPLES FROM THE WEB sentences
SENTENCES_DIV_CLASS = 'css-79elbk e1dkhfa64'


def get_soup(word):
    """
    :param word: the word to look up in thesaurus.com
    :return: a BeautifulSoup object that contains the html of the word web page
    """
    result = requests.get(f'https://www.thesaurus.com/browse/{word}')
    return BeautifulSoup(result.text, 'html.parser')


def get_synonyms(soup, synonyms_num):
    """
    :param soup: the BeautifulSoup object of the word web page
    :param synonyms_num: the desired number of synonyms
    :return: a list that contains the first synonyms_num synonyms
    """
    synonyms_divs = soup.find_all('div', {'class': SYNONYMS_DIV_CLASS})
    synonyms = []
    for div in synonyms_divs[:synonyms_num]:
        # get the text of the first child node
        synonyms.append(div.findChild().text)
    return synonyms


def get_sentences(soup, sentences_num):
    """
    :param soup: the BeautifulSoup object of the word web page
    :param sentences_num: the desired number of sentences
    :return: a list that contains the first sentences_num sentences
    """
    sentences_div = soup.find('div', {'class': SENTENCES_DIV_CLASS})
    if sentences_div is None:
        return []
    p_tags = sentences_div.find_all('p')
    return [p.text for p in p_tags[:sentences_num]]


def save_excel_file(words_data):
    """
    Saves the words_data to a styled excel sheet
    :param words_data: [[word, synonyms, sentences], ....]
    :return: None
    """
    workbook = Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = 'word'
    for i in range(SYNONYMS_NUM):
        column_index = i + 2
        sheet.cell(row=1, column=column_index).value = f'synonym {i+1}'
        sheet.column_dimensions[get_column_letter(
            column_index)].width = FONT_SIZE + 5
    for i in range(SENTENCES_NUM):
        column_index = i + 2 + SYNONYMS_NUM
        sheet.cell(row=1, column=column_index).value = f'sentence {i+1}'
        sheet.column_dimensions[get_column_letter(
            column_index)].width = FONT_SIZE * 7

    row_index = 2  # starts at 1 + header size (1)
    for word, synonyms, sentences in words_data:
        sheet.cell(row=row_index, column=1).value = word
        for i, synonym in enumerate(synonyms):
            column_index = i + 2
            sheet.cell(row=row_index, column=column_index).value = synonym
        for i, sentence in enumerate(sentences):
            column_index = i + 2 + SYNONYMS_NUM
            sheet.cell(row=row_index, column=column_index).value = sentence
        row_index += 1

    font = Font(size=FONT_SIZE)
    sheet.column_dimensions['A'].width = FONT_SIZE * 2.5
    sheet.row_dimensions[1].height = FONT_SIZE * 2
    sheet.freeze_panes = 'B2'

    # apply the font to all the cells in the seet
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font

    # style template for the header row
    header = NamedStyle(name="header")
    header.font = Font(size=FONT_SIZE, bold=True)
    header.border = Border(bottom=Side(border_style="double"))
    header.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet[1]:
        cell.style = header

    workbook.save(filename=OUTPUT_FILENAME)
    print(f'The data was saved successfully to {OUTPUT_FILENAME}')


def main():
    """
    Iterates over the words in WORDS_LIST_PATH,
    and for each words looks up synonyms and sentences in thesaurus.com
    finally save all the data to an excel sheet
    """
    words_data = []

    if not os.path.exists(WORDS_LIST_PATH):
        print('WORDS_LIST_PATH is incorrect!')
        print('create a file with your words and update its path.')
        return

    with open(WORDS_LIST_PATH, 'r') as file:
        for i, line in enumerate(file):
            line = line.strip()
            if line == '':
                continue

            pattern = f"[{''.join(WORDS_SEPARATORS)}]+"
            words = [word.strip() for word in re.split(pattern, line)]

            synonyms = []
            sentences = []
            for j, word in enumerate(words):
                soup = get_soup(word)
                synonyms += get_synonyms(soup,
                                         round(SYNONYMS_NUM / len(words)))
                if j == 0:  # only for the first word in the line
                    sentences += get_sentences(soup, SENTENCES_NUM)

            words_data.append([line, synonyms, sentences])
            print(f"{i+1}) {line} - {', '.join(synonyms)} - {', '.join(sentences)}")

    save_excel_file(words_data)


if __name__ == '__main__':
    main()
